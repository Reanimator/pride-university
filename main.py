from openpyxl import Workbook
import pathlib
from pathlib import Path


def make_report_about_top3(students_avg_scores):
    """
    Функция для поиска лучших студентов и записи в Excel
    :param students_avg_scores: словарь со студентами и их средними оценками
    :return: ссылка на Excel файл
    """
    top_students = sorted(students_avg_scores.items(), key=lambda student: student[1], reverse=True)[0:3]
    wb = Workbook()
    ws = wb.active
    ws['B2'] = 'Список лучших студентов'
    ws['B3'] = 'Студент'
    ws['C3'] = 'Ср. оценка'
    for index, student in enumerate(top_students):
        ws[f'B{index+4}'] = student[0]
        ws[f'C{index+4}'] = student[1]
    wb.save('top_students.xlsx')
    dir_path = pathlib.Path.cwd()
    path = Path(dir_path, 'top_students.xlsx')
    return path


if __name__ == "__main__":
    students_avg_scores = {'Max': 4.964, 'Eric': 4.962, 'Peter': 4.923, 'Mark': 4.957, 'Julie': 4.95, 'Jimmy': 4.973,
                           'Felix': 4.937, 'Vasya': 4.911, 'Don': 4.936, 'Zoi': 4.937}
    print(make_report_about_top3(students_avg_scores))
